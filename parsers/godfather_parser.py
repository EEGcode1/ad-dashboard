"""
Godfather the Offer — Bangbang Games collab.

Source: SGI_Reporting@lnw.com
Subject: "A Game for Yesterday - Godfather the Offer"
Attachment: small .xlsx (~30KB) scoped to a single game.

IMPORTANT: this is NOT the main LNW daily revenue. It's a separate
game-specific report for the Godfather the Offer title, a collab
with Bangbang Games (contact: Kerri). It must be filed under its own
line item and cross-checked against Kerri's monthly statements.

Output:
  {
    "source": "godfather_collab",
    "partner": "bangbang",
    "date": "YYYY-MM-DD",
    "wager_eur": float,
    "ggr_eur": float,
    "rtp": float or None,
    # rev share for this collab is negotiated separately; do NOT auto-apply 4%.
  }
"""

from __future__ import annotations

import io
import re
from dataclasses import dataclass
from datetime import datetime
from typing import Optional

from openpyxl import load_workbook

from . import gmail_client

GODFATHER_SENDER = "SGI_Reporting@lnw.com"
GODFATHER_SUBJECT_MATCH = "A Game for Yesterday - Godfather the Offer"


@dataclass
class GodfatherRow:
    source: str
    partner: str
    date: str
    wager_eur: float
    ggr_eur: float
    rtp: Optional[float]

    def to_dict(self) -> dict:
        return {
            "source": self.source,
            "partner": self.partner,
            "date": self.date,
            "wager_eur": round(self.wager_eur, 2),
            "ggr_eur": round(self.ggr_eur, 2),
            "rtp": self.rtp,
        }


def _find_date_in_sheet(ws) -> Optional[str]:
    for row in ws.iter_rows(min_row=1, max_row=8, values_only=True):
        for v in row:
            if isinstance(v, datetime):
                return v.date().isoformat()
            if isinstance(v, str):
                m = re.search(r"(\d{4})[-/](\d{1,2})[-/](\d{1,2})", v)
                if m:
                    y, mo, d = m.groups()
                    return f"{int(y):04d}-{int(mo):02d}-{int(d):02d}"
                m = re.search(r"(\d{1,2})[-/](\d{1,2})[-/](\d{4})", v)
                if m:
                    d, mo, y = m.groups()
                    return f"{int(y):04d}-{int(mo):02d}-{int(d):02d}"
    return None


def parse_xlsx_bytes(xlsx_bytes: bytes) -> GodfatherRow:
    wb = load_workbook(io.BytesIO(xlsx_bytes), data_only=True, read_only=True)
    ws = wb.active
    report_date = _find_date_in_sheet(ws) or datetime.utcnow().date().isoformat()

    wager = 0.0
    ggr = 0.0
    rtp = None

    # Small file — scan all rows for labeled values.
    for row in ws.iter_rows(values_only=True):
        if not row:
            continue
        label = str(row[0] or "").strip().lower()
        if not label:
            continue
        value_cells = [c for c in row[1:] if c not in (None, "")]
        if not value_cells:
            continue
        first = value_cells[0]
        try:
            num = float(first)
        except (TypeError, ValueError):
            continue
        if "wager" in label or "stake" in label or "turnover" in label:
            wager = num
        elif "ggr" in label or "gross" in label:
            ggr = num
        elif "rtp" in label or "payout" in label:
            rtp = num

    return GodfatherRow(
        source="godfather_collab",
        partner="bangbang",
        date=report_date,
        wager_eur=wager,
        ggr_eur=ggr,
        rtp=rtp,
    )


def fetch_latest(service=None, lookback_days: int = 3) -> Optional[GodfatherRow]:
    if service is None:
        service = gmail_client.get_service()
    query = (
        f'from:{GODFATHER_SENDER} subject:"{GODFATHER_SUBJECT_MATCH}" '
        f'has:attachment newer_than:{lookback_days}d'
    )
    ids = gmail_client.search_messages(service, query, max_results=3)
    if not ids:
        return None
    msg = gmail_client.get_message(service, ids[0])
    xlsx = next((a for a in msg.attachments if a["filename"].lower().endswith(".xlsx")), None)
    if not xlsx:
        return None
    data = gmail_client.download_attachment(service, msg.id, xlsx["attachment_id"])
    return parse_xlsx_bytes(data)


if __name__ == "__main__":
    row = fetch_latest()
    if row:
        print(f"Godfather/Bangbang {row.date}: wager €{row.wager_eur:,.2f} GGR €{row.ggr_eur:,.2f}")
    else:
        print("No Godfather report in lookback window.")
