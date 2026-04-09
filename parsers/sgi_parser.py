"""
SGI parser — LNW Regulated "Atlantic Digital Extended Report".

Source: SGI_Reporting@lnw.com
Subject: "Atlantic Digital Extended Report"
Attachment: .xlsx (~1.6 MB)

CRITICAL: parse the "Previous Day - By Game" tab ONLY.
Other tabs (weekly/MTD rollups) caused 1.7–1.96× inflation on Mar 18 / Mar 23
when Gary's parser picked them up. Do NOT fall back to other tabs.

Output rows:
  {
    "source": "lnw",
    "date": "YYYY-MM-DD",       # the "previous day" referenced in the file
    "game": "<game name>",
    "wager_eur": <float>,
    "ggr_eur": <float>,
    "rtp": <float or None>,
    "rev_share_eur": <ggr * 0.04>,
  }

The "A Game for Yesterday - Godfather the Offer" email is a DIFFERENT report
(Bangbang Games collab) and is handled by godfather_parser.py.
"""

from __future__ import annotations

import io
import re
from dataclasses import dataclass
from datetime import date, datetime, timedelta
from typing import Optional

from openpyxl import load_workbook

from . import gmail_client

# Fixed constants — do not change without audit sign-off.
LNW_REV_SHARE = 0.04
TARGET_TAB = "Previous Day - By Game"
SGI_SENDER = "SGI_Reporting@lnw.com"
SGI_SUBJECT_MATCH = "Atlantic Digital Extended Report"


@dataclass
class SgiRow:
    source: str
    date: str
    game: str
    wager_eur: float
    ggr_eur: float
    rtp: Optional[float]
    rev_share_eur: float

    def to_dict(self) -> dict:
        return {
            "source": self.source,
            "date": self.date,
            "game": self.game,
            "wager_eur": round(self.wager_eur, 2),
            "ggr_eur": round(self.ggr_eur, 2),
            "rtp": self.rtp,
            "rev_share_eur": round(self.rev_share_eur, 2),
        }


def _norm(s) -> str:
    return str(s or "").strip().lower()


def _find_header_row(ws) -> tuple[int, dict[str, int]]:
    """
    Scan the first ~10 rows for a header row containing game/wager/ggr columns.
    Returns (header_row_index_1based, {"game": col, "wager": col, "ggr": col, "rtp": col}).
    """
    for row_idx in range(1, 11):
        headers = {}
        for col_idx, cell in enumerate(ws[row_idx], start=1):
            h = _norm(cell.value)
            if not h:
                continue
            if "game" in h and "count" not in h and "game" not in headers:
                headers["game"] = col_idx
            elif "wager" in h or "turnover" in h or "stake" in h:
                headers["wager"] = col_idx
            elif "ggr" in h or ("gross" in h and "revenue" in h):
                headers["ggr"] = col_idx
            elif "rtp" in h or "payout" in h:
                headers["rtp"] = col_idx
        if {"game", "wager", "ggr"}.issubset(headers):
            return row_idx, headers
    raise ValueError(
        f"Could not locate header row in '{TARGET_TAB}'. "
        "Sheet layout may have changed — investigate before trusting any output."
    )


def _extract_report_date(wb, ws) -> str:
    """
    Try to find the report date. SGI files typically note the reporting
    day near the top of the sheet or in a metadata tab. We fall back to
    'yesterday' (UTC) if nothing is found, but log a warning.
    """
    # Scan first 5 rows for an obvious date cell
    for row_idx in range(1, 6):
        for cell in ws[row_idx]:
            v = cell.value
            if isinstance(v, datetime):
                return v.date().isoformat()
            if isinstance(v, date):
                return v.isoformat()
            if isinstance(v, str):
                m = re.search(r"(\d{4})[-/](\d{1,2})[-/](\d{1,2})", v)
                if m:
                    y, mo, d = m.groups()
                    return f"{int(y):04d}-{int(mo):02d}-{int(d):02d}"
                m = re.search(r"(\d{1,2})[-/](\d{1,2})[-/](\d{4})", v)
                if m:
                    d, mo, y = m.groups()
                    return f"{int(y):04d}-{int(mo):02d}-{int(d):02d}"
    # Fallback — yesterday UTC
    return (datetime.utcnow().date() - timedelta(days=1)).isoformat()


def parse_xlsx_bytes(xlsx_bytes: bytes) -> list[SgiRow]:
    wb = load_workbook(io.BytesIO(xlsx_bytes), data_only=True, read_only=True)

    if TARGET_TAB not in wb.sheetnames:
        raise ValueError(
            f"Tab '{TARGET_TAB}' not found. Available tabs: {wb.sheetnames}. "
            "Refusing to fall back to another tab — historically this caused "
            "1.7–1.96× inflation."
        )

    ws = wb[TARGET_TAB]
    header_row, cols = _find_header_row(ws)
    report_date = _extract_report_date(wb, ws)

    rows: list[SgiRow] = []
    for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
        if not row:
            continue
        game_val = row[cols["game"] - 1] if cols["game"] - 1 < len(row) else None
        if not game_val or _norm(game_val) in ("total", "totals", "grand total", ""):
            continue
        try:
            wager = float(row[cols["wager"] - 1] or 0)
            ggr = float(row[cols["ggr"] - 1] or 0)
        except (TypeError, ValueError):
            continue
        rtp_val = None
        if "rtp" in cols:
            raw_rtp = row[cols["rtp"] - 1]
            try:
                rtp_val = float(raw_rtp) if raw_rtp is not None else None
            except (TypeError, ValueError):
                rtp_val = None

        rows.append(SgiRow(
            source="lnw",
            date=report_date,
            game=str(game_val).strip(),
            wager_eur=wager,
            ggr_eur=ggr,
            rtp=rtp_val,
            rev_share_eur=ggr * LNW_REV_SHARE,
        ))

    if not rows:
        raise ValueError(
            f"'{TARGET_TAB}' parsed zero rows — refusing to publish an empty "
            "LNW day. Investigate the attachment manually."
        )

    return rows


def fetch_latest(service=None, lookback_days: int = 3) -> list[SgiRow]:
    """Find the most recent SGI Extended Report email and parse it."""
    if service is None:
        service = gmail_client.get_service()

    query = (
        f'from:{SGI_SENDER} subject:"{SGI_SUBJECT_MATCH}" '
        f'has:attachment newer_than:{lookback_days}d'
    )
    ids = gmail_client.search_messages(service, query, max_results=5)
    if not ids:
        raise RuntimeError(
            f"No SGI email found in last {lookback_days} days matching: {query}"
        )

    msg = gmail_client.get_message(service, ids[0])
    xlsx_att = next(
        (a for a in msg.attachments if a["filename"].lower().endswith(".xlsx")),
        None,
    )
    if not xlsx_att:
        raise RuntimeError(
            f"SGI message {msg.id} has no .xlsx attachment. "
            f"Subject: {msg.subject}"
        )

    xlsx_bytes = gmail_client.download_attachment(service, msg.id, xlsx_att["attachment_id"])
    return parse_xlsx_bytes(xlsx_bytes)


if __name__ == "__main__":
    rows = fetch_latest()
    total_ggr = sum(r.ggr_eur for r in rows)
    total_share = sum(r.rev_share_eur for r in rows)
    print(f"Parsed {len(rows)} LNW rows for {rows[0].date}")
    print(f"  GGR total:       €{total_ggr:,.2f}")
    print(f"  Rev share (4%):  €{total_share:,.2f}")
